import openpyxl
import COCOZZA_connectDb, datetime
from COCOZZA_logs import PRINT, SALVA_LOG



# Apri il file Excel
workbook = openpyxl.load_workbook('./fatture_sinistri.xlsx')

# Seleziona il foglio di lavoro "pagina 1"
worksheet = workbook['FATTURE']

# Leggi i titoli della tabella dalla riga 1, colonne A a W
titles = []
for column in range(1, 11):
    cell_value = worksheet.cell(row=1, column=column).value
    titles.append(cell_value)

# Leggi i dati delle righe successive, dalle celle A2 a Wn (n = numero di righe)
data = []
for row in range(2, worksheet.max_row + 1):
    row_data = {}
    for column in range(1, 11):
        cell_value = worksheet.cell(row=row, column=column).value
        row_data[titles[column - 1]] = cell_value
    data.append(row_data)

# Chiudi il file Excel
workbook.close()

# Fai qualcosa con i dati letti...
print(titles)

# COCOZZA_connectDb.start()


response1 = input("Hai creato gli utenti 'Petit', 'Cardone Gerardo', 'Milkman'?")
response2 = input("Hai messo nel file excel IDTRATTENUTE in ordine crescente?")



for record in data:
    if record['IDFATTURA'] <= 0:
        continue
    PRINT("\nSTAMPO RECORD\n\n")
    PRINT(record)
    sinistri_id = record['ID_ODOO']
    datetime_trattenuta = record['DATA']
    fattura = record['FATTURA']
    valore = record['IMPORTO']
    descrizione = record['NOTE']
    
    petit_id = COCOZZA_connectDb.sqlSearch('res.partner',['name', 'ilike', 'Petit'],'')[0]['id']

    
    PRINT(sinistri_id)    
    PRINT(datetime_trattenuta) 
    PRINT(fattura)       
    PRINT(valore)
    PRINT(descrizione)      
    PRINT(petit_id)
    
    
        
    # Se il datime è vuoto imposto la data vuota
    if datetime_trattenuta == None:
        data_trattenuta = ""
    else:
        # Trasformo il datetime in data
        dt = datetime.datetime.strptime(str(datetime_trattenuta), "%Y-%m-%d %H:%M:%S")
        data_trattenuta = dt.date()
        PRINT(data_trattenuta)



    PRINT("Procedo con la  creazione della fatturazione\n")
    PRINT(f"fleet_vehicle_log_service_id = {sinistri_id}")
    PRINT(f"date = {str(data_trattenuta)}")
    PRINT(f"partner_id = {petit_id}")
    PRINT(f"invoice_ref = {fattura}")
    PRINT(f"reparation_value = {valore}")
    PRINT(f"description = {descrizione}")

    # Differenzio la query di creazione in base a se esiste la data o meno
    if datetime_trattenuta != None:
        # Creazione record
        COCOZZA_connectDb.sqlCreate('reparation.reparation', {
            'fleet_vehicle_log_service_id': sinistri_id,
            'date': str(data_trattenuta),
            'partner_id': petit_id,
            'invoice_ref': fattura,
            'reparation_value': valore,
            'description': descrizione,

        })
        PRINT("Creazione eseguita")
    else:
        # Creo una queri senza data in quanto non inserita nel file excel
        COCOZZA_connectDb.sqlCreate('reparation.reparation', {
            'fleet_vehicle_log_service_id': sinistri_id,
            'partner_id': petit_id,
            'invoice_ref': fattura,
            'reparation_value': valore,
            'description': descrizione,
        })
        PRINT("Creazione eseguita")

PRINT("Importato tutti i dati.")