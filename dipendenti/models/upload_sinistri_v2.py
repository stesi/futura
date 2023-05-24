import openpyxl
import COCOZZA_connectDb
from COCOZZA_logs import PRINT, SALVA_LOG
import threading

# Apri il file Excel
workbook = openpyxl.load_workbook('./sinistri.xlsx')

# Seleziona il foglio di lavoro "SINISTRI"
worksheet = workbook['SINISTRI']

# Leggi i titoli della tabella dalla riga 1, colonne A a W
titles = []
for column in range(1, 28):
    cell_value = worksheet.cell(row=1, column=column).value
    titles.append(cell_value)

# Leggi i dati delle righe successive, dalle celle A2 a Wn (n = numero di righe)
data = []
for row in range(2, worksheet.max_row + 1):
    row_data = {}
    for column in range(1, 28):
        cell_value = worksheet.cell(row=row, column=column).value
        row_data[titles[column - 1]] = cell_value
    data.append(row_data)

# Chiudi il file Excel
workbook.close()

# Fai qualcosa con i dati letti...
print(titles)

# COCOZZA_connectDb.start()

# Funzione per elaborare un singolo record
def process_record(record, semaphore):
    if record['IDSINISTRI'] <= 0:
        return
    
    PRINT("\nSTAMPO RECORD\n\n")
    PRINT(record)
    
    # Estrai i valori dai record
    id_sinistro = record['IDSINISTRI']
    intestatario = record['INTESTATARIO']
    data = record['DATA']
    valore_danno = record['VALORE DANNO']
    targa = record['TARGA']
    soggetto = record['AUTISTA']
    tipo_danno = record['TIPO DANNO']
    competenza = record['COMPETENZA']
    descrizione = record['DESCRIZIONE']
    
    
    PRINT(id_sinistro)
    PRINT(intestatario)
    PRINT(data)
    PRINT(valore_danno)
    PRINT(targa)
    PRINT(soggetto)
    PRINT(tipo_danno)
    PRINT(competenza)
    PRINT(descrizione)
    
    # Esegui le operazioni desiderate con i valori estratti
    
    # Esegui le operazioni di ricerca nel database
    res_partner_id = COCOZZA_connectDb.sqlSearch('res.partner', ['name', '=', soggetto])
    vehicle_id = COCOZZA_connectDb.sqlSearch('fleet.vehicle', ['license_plate', '=', targa])
    
    # Esegui i controlli e le operazioni aggiuntive
    if res_partner_id == []:
        SALVA_LOG("sinistri_saltati_dipendenti.txt", f"Il dipendente {soggetto} non esiste nei res.partner\n")
        return
    if vehicle_id == []:
        SALVA_LOG("sinistri_saltati_targa.txt", f"Il veicolo con targa {targa} non esiste nel fleet.vehicle\n")
        return
    
    # Esegui le operazioni di creazione del record nel database
    COCOZZA_connectDb.sqlCreate('fleet.vehicle.log.services', {
        'description': "id vecchio gestionale " + str(id_sinistro),
        'service_type_id': service_type_log_id[0]['id'],
        'date': str(data),  # DA INSERIRE L'ORA!!!
        'amount': valore_danno,
        'vehicle_id': vehicle_id[0]['id'],
        'purchaser_id': res_partner_id[0]['id'],
        'responsibility': responsability,
        'repair_mode': modalita_riparazione,
        'notes': descrizione,
        'state': "done",
    })
    
    PRINT("Creazione eseguita")

# Creazione del semaforo con limite di 3 thread
max_threads = 3
semaphore = threading.BoundedSemaphore(max_threads)

# Creazione di thread per ogni record
threads = []
for record in data:
    # Acquisisci il semaforo prima di avviare il thread
    semaphore.acquire()
    
    # Crea e avvia il thread
    thread = threading.Thread(target=process_record, args=(record, semaphore))
    threads.append(thread)
    thread.start()

# Attendi il completamento di tutti i thread
for thread in threads:
    thread.join()

# Rilascia il semaforo
semaphore.release()
