import openpyxl
import COCOZZA_connectDb
from COCOZZA_logs import PRINT, SALVA_LOG



# Apri il file Excel
workbook = openpyxl.load_workbook('./sinistri.xlsx')

# Seleziona il foglio di lavoro "pagina 1"
worksheet = workbook['SINISTRI']

# Leggi i titoli della tabella dalla riga 1, colonne A a W
titles = []
for column in range(1, 28):
    cell_value = worksheet.cell(row=1, column=column).value
    titles.append(cell_value)

# Leggi i dati delle righe successive, dalle celle A2 a Wn (n = numero di righe)
data = []
for row in range(2, worksheet.max_row + 1):
    row_data = {}
    for column in range(1, 28):
        cell_value = worksheet.cell(row=row, column=column).value
        row_data[titles[column - 1]] = cell_value
    data.append(row_data)

# Chiudi il file Excel
workbook.close()

# Fai qualcosa con i dati letti...
print(titles)


# COCOZZA_connectDb.start()



for record in data:
    PRINT("\nSTAMPO RECORD\n\n")
    PRINT(record)
    
    id_sinistro = record['IDSINISTRI']
    intestatario = record['INTESTATARIO']
    data = record['DATA']
    valore_danno = record['VALORE DANNO']
    targa = record['TARGA']
    soggetto = record['AUTISTA']
    tipo_danno = record['TIPO DANNO']
    competenza = record['COMPETENZA']
    descrizione = record['DESCRIZIONE']

    
    

    
    print(id_sinistro)    
    print(intestatario)    
    print(data)    
    print(valore_danno)    
    print(targa)    
    print(soggetto)    
    print(tipo_danno)    
    print(competenza)    
    print(descrizione)    
    
    
    print(COCOZZA_connectDb.sqlSearch('res.partner',['name', 'ilike', soggetto]))
    


    # Recupero ID
    service_type_log_id = COCOZZA_connectDb.sqlSearch('fleet.service.type', ['name', '=', 'Sinistro'])
    vehicle_id = COCOZZA_connectDb.sqlSearch('fleet.vehicle', ['license_plate', '=', targa])
    res_partner_id = COCOZZA_connectDb.sqlSearch('res.partner', ['name', '=', soggetto])
    
    
    # Controlli anti interruzione con salvataggio dei record saltati
    if res_partner_id == []:
        SALVA_LOG("sinistri_saltati_dipendenti.txt",f"Il dipendente {soggetto} non esiste nei res.partner\n")
        continue
    if vehicle_id == []:
        SALVA_LOG("sinistri_saltati_targa.txt",f"Il veicolo con targa {targa} non esiste nel fleet.vehicle\n")
        continue
    
    # Setto il valore di Responsabilità alla variabile
    if tipo_danno == "PROPRI":
        responsability = "byself"
    elif tipo_danno == "TERZI":
        responsability = "third"
    elif tipo_danno == "PROPRI + TERZI":
        responsability = "byself_third"
    elif tipo_danno == "PROPRI DA IGNOTI":
        responsability = "unknown"



    # Setto il valore di modalità riparazione alla variabile
    if competenza == "ATTESA AUTORIPARAZIONE" or competenza == "AUTORIPARAZIONE":
        modalita_riparazione = "internal"
    else:
        modalita_riparazione = "external"


    PRINT("Procedo con la creazione del seguente sinistro:\n")
    PRINT("id vecchio gestionale " + str(id_sinistro))
    PRINT(f"date = {service_type_log_id[0]['id']}")
    PRINT(f"amount = {valore_danno}")
    PRINT(f"vehicle_id = {vehicle_id[0]['id']}")
    PRINT(f"Il soggetto è {soggetto}")
    PRINT(f"purchaser_id = {res_partner_id[0]['id']}")
    PRINT(f"Tipo di danno = {responsability}")
    PRINT(f"modalità riparazione = {modalita_riparazione}")
    PRINT(f"notes = {descrizione}")
    
    
    
    # Creazine record
    COCOZZA_connectDb.sqlCreate('fleet.vehicle.log.services', {
        'description': "id vecchio gestionale " + str(id_sinistro),
        'service_type_id': service_type_log_id[0]['id'],
        'date': str(data),
        'amount': valore_danno,
        'vehicle_id': vehicle_id[0]['id'],
        'purchaser_id': res_partner_id[0]['id'],
        'responsibility': responsability,
        'repair_mode': modalita_riparazione,
        'notes': descrizione,
        'state': "done",
    })
    PRINT("Creazione eseguita")
    