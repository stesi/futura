import openpyxl
import COCOZZA_connectDb
from COCOZZA_logs import PRINT, SALVA_LOG



# Apri il file Excel
workbook = openpyxl.load_workbook('./multe.xlsx')

# Seleziona il foglio di lavoro "pagina 1"
worksheet = workbook['MULTE']

# Leggi i titoli della tabella dalla riga 1, colonne A a W
titles = []
for column in range(1, 23):
    cell_value = worksheet.cell(row=1, column=column).value
    titles.append(cell_value)

# Leggi i dati delle righe successive, dalle celle A2 a Wn (n = numero di righe)
data = []
for row in range(2, worksheet.max_row + 1):
    row_data = {}
    for column in range(1, 23):
        cell_value = worksheet.cell(row=row, column=column).value
        row_data[titles[column - 1]] = cell_value
    data.append(row_data)

# Chiudi il file Excel
workbook.close()

# Fai qualcosa con i dati letti...
print(titles)


# COCOZZA_connectDb.start()

# sconosciuto1 = input("Hai creato il veicolo del produttore 'Sconosciuto', il modello 'Sconosciuto' e con la targa 'Sconosciuto'???")
# sconosciuto2 = input("Hai settato sul file excel le targe vuote con 'Sconosciuto'?\nHai settato sul file excel i soggetti vuoti con 'Sconosciuto'?")

# if sconosciuto1.lower() != "si" and sconosciuto2.lower() != "si":
#     quit()

for record in data:
    if record['SOGGETTO'] != 'BEN KRAM ALAAEDDINE':
        continue
    PRINT("\nSTAMPO RECORD\n\n")
    PRINT(record)
    id_multa = record['IDMULTA']
    n_verbale = record['N° VERBALE']
    intestatario = record['INTESTATARIO']
    data_verbale = record['DATA VERBALE']
    ora_verbale = record['ORA']
    luogo = record['LUOGO']
    comune = record['CITTA\'']
    targa = record['TARGA']
    infrazione = record['INFRAZIONE']
    importo = record['IMPORTO']
    competenza = record['COMPETENZA']
    decurtazione_punti = record['DECURTAZIONE PUNTI']
    soggetto = record['SOGGETTO']
    note = record['INFRAZIONE']

    # Creazione Datetime
    ore, minuti = ora_verbale.split(':')
    datetime = data_verbale.replace(hour=int(ore), minute=int(minuti))
    
    print(id_multa)    
    print(n_verbale)    
    print(intestatario)    
    print(datetime)    
    print(ora_verbale)    
    print(luogo)    
    print(comune)    
    print(targa)    
    print(infrazione)    
    print(importo)    
    print(competenza)    
    print(decurtazione_punti)    
    print(soggetto) 
    print(COCOZZA_connectDb.sqlSearch('res.partner',['name', 'ilike', soggetto]))
    


    # Recupero ID
    service_type_log_id = COCOZZA_connectDb.sqlSearch('fleet.service.type', ['name', '=', 'Multa'])
    vehicle_id = COCOZZA_connectDb.sqlSearch('fleet.vehicle', ['license_plate', '=', targa])
    res_partner_id = COCOZZA_connectDb.sqlSearch('res.partner', ['name', '=', soggetto])
    
    
    # Controlli anti interruzione con salvataggio dei record saltati
    if res_partner_id == []:
        SALVA_LOG("multe_saltate_dipendenti.txt", f"Il dipendente {soggetto} non esiste nei res.partner\n")
        continue
    if vehicle_id == []:
        SALVA_LOG("multe_saltate_veicoli.txt", f"Il veicolo con targa {targa} non esiste nel fleet.vehicle\n")
        continue
    
    
    
    # Sistemazione variabili
    if decurtazione_punti == None:
        punti = 0
        print("punto è settato a ZERO")
    else:
        punti = decurtazione_punti
        print("Punti è settato con il valore dell'Excel")

    
    #Saltare per problemi utenti
    
    saltare_soggetto = ['SENATORE ANTONIO', 'ZORDAN MASSIMO', "", "LOMBARDI SIMONE", "CAVE MICHELE", "DELLAL KARIM", "VECCHIETTI RUGGERO", "BELLASSAI EMILIANO"]
    saltare_targa = ['FB870DY', "EV096MK", "EY096MK"]

    if soggetto == "":
        print("")
    PRINT("Procedo con la creazione della seguente multa:\n")
    PRINT("N° verbale " + str(n_verbale) + " - id vecchio gestionale " + str(id_multa))
    PRINT(f"date = {datetime}")
    PRINT(f"amount = {importo}")
    PRINT(f"vehicle_id = {vehicle_id[0]['id']}")
    PRINT(f"Il soggetto è {soggetto}")
    PRINT(f"purchaser_id = {res_partner_id[0]['id']}")
    PRINT(f"deduction_point = {punti}")
    PRINT(f"notes = {note}")
    
    
    
# Creazione record
    COCOZZA_connectDb.sqlCreate('fleet.vehicle.log.services', {
        'description': "N° verbale " + str(n_verbale) + " - id vecchio gestionale " + str(id_multa),
        'service_type_id': service_type_log_id[0]['id'],
        'date': str(datetime),
        'amount': importo,
        'vehicle_id': vehicle_id[0]['id'],
        'purchaser_id': res_partner_id[0]['id'],
        'deduction_point': punti,
        'notes': note,
        'state': "done",
    })
    PRINT("Creazione eseguita")

