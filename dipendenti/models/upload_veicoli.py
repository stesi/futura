import openpyxl
import COCOZZA_connectDb
from COCOZZA_logs import PRINT



# Apri il file Excel
workbook = openpyxl.load_workbook('./file.xlsx')

# Seleziona il foglio di lavoro "pagina 1"
worksheet = workbook['pagina 1']

# Leggi i titoli della tabella dalla riga 1, colonne A a G
titles = []
for column in range(1, 8):
    cell_value = worksheet.cell(row=1, column=column).value
    titles.append(cell_value)

# Leggi i dati delle righe successive, dalle celle A2 a Gn (n = numero di righe)
data = []
for row in range(2, worksheet.max_row + 1):
    row_data = {}
    for column in range(1, 8):
        cell_value = worksheet.cell(row=row, column=column).value
        row_data[titles[column - 1]] = cell_value
    data.append(row_data)

# Chiudi il file Excel
workbook.close()

# Fai qualcosa con i dati letti...
print(titles)
print(data[1]['STATO'])


COCOZZA_connectDb.start()



for record in data:
    PRINT("STAMPO RECORD")
    PRINT(record)
    stato = record['STATO']
    targa = record['TARGA']
    cdc = record['CDC']
    brand = record['BRAND']
    modello = record['MODELLO']
    categoria = record['TIPO']
    euro = record['EURO']
    
    # Verifico se il brand, il modello e la targa sono già inseriti nel db
    is_targa = COCOZZA_connectDb.sqlSearch('fleet.vehicle', ['name', '=', targa])
    is_cdc = COCOZZA_connectDb.sqlSearch('res.partner', ['name', '=', cdc])
    
    # print("FACCIO UN POO' DI PRINT")
    # print(is_brand)
    # print(is_modello)
    # print(is_categoria)
    # print(is_targa)
    # print(is_cdc)

    if targa == None:
        break

    # Verifico se la targa è già inserita
    PRINT("Controllo se la targa è inserita")
    controllo_targa = COCOZZA_connectDb.sqlSearch('fleet.vehicle',['license_plate', '=', targa])
    print(controllo_targa)
    if controllo_targa == []:
        print("Entro nella if inerente al veicolo non esistente")
        PRINT(f"La targa {targa} non è esistente.")
        # Verifico se il brand esiste
        is_brand = COCOZZA_connectDb.sqlSearch('fleet.vehicle.model.brand', ['name', '=', brand.lower().capitalize()])
        if is_brand == []:
            PRINT(f"il Brand {brand} non esiste. Procedo con la creazione.")
            is_brand = COCOZZA_connectDb.sqlCreate('fleet.vehicle.model.brand',{'name', '=', brand.lower().capitalize()})
            PRINT(f"Brand creato.")
        PRINT("Stampo is-brand")
        PRINT(is_brand)
        PRINT("Verifica del brand terminata")
        
        # Verifico se la categoria esiste
        is_categoria = COCOZZA_connectDb.sqlSearch('fleet.vehicle.model.category', ['name', '=', categoria.lower().capitalize()])
        if is_categoria == []:
            PRINT(f"La categoria {categoria} non esiste. Procedo con la creazione della categoria.")
            is_categoria = COCOZZA_connectDb.sqlCreate('fleet.vehicle.model.category', {'name': categoria.lower().capitalize()})
            PRINT(f"Categoria {categoria} creata con successo")
        PRINT("Stampo is-categoria")
        PRINT(is_categoria)
        PRINT("Verifica della categoria terminata")

        # Verifico che il modello esista
        is_modello = COCOZZA_connectDb.sqlSearchMultiple('fleet.vehicle.model', [('name', '=', modello.lower().capitalize()), ('category_id', '=', is_categoria[0]['id']),('brand_id', '=', is_brand[0]['id'])],{'fields': ['id'], 'limit': 1})
        print(f"Stampo modello      : {is_modello}")
        if is_modello == []:
            PRINT(f"Il modello {modello} non esiste. Procedo con la creazione del modello.")
            is_modello = COCOZZA_connectDb.sqlCreate('fleet.vehicle.model', {'name': modello.lower().capitalize(),
                                                                'brand_id': is_brand[0]['id'],
                                                                'category_id': is_categoria[0]['id']})
            modello_id = is_modello[0]
            PRINT("Modello creato correttamente con id")
        else:
            modello_id = is_modello[0]['id']
        print("CI PROVOO")
        PRINT('Stampo is_modello')
        PRINT(is_modello)
        PRINT('Stampo modello_id')
        PRINT(modello_id)
        PRINT("Verifica del modello terminata")

        # Creo il mezzo
        PRINT(f"Procedo con la creazione del veicolo con targa {targa}")
        print(f"Ecco is_brand = {is_brand[0]['id']}")
        print(f"Ecco targa = {targa}")
        print(f"Ecco id_cdc = {is_cdc[0]['id']}")
        print(f"Ecco modello_id = {modello_id}")
        COCOZZA_connectDb.sqlCreate("fleet.vehicle", {
                                                    'brand_id': is_brand[0]['id'],
                                                    'license_plate': targa,
                                                    'organization_id': is_cdc[0]['id'],
                                                    'model_id': modello_id,
                                                    'euro': euro,
                                                    'state_vehicle': stato
                                                    })
        PRINT("Veicolo aggiunto!")
    
    else:
        PRINT(f"Veicolo con targa {targa} già esistente.")
PRINT("L'IMPORTAZIONE E' FINALMENTE TERMINATA!!!")