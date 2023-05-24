import openpyxl
import COCOZZA_connectDb
from COCOZZA_logs import PRINT



# Apri il file Excel
workbook = openpyxl.load_workbook('./targhe_mancanti.xlsx')

# Seleziona il foglio di lavoro "pagina 1"
worksheet = workbook['Foglio1']

# Leggi i titoli della tabella dalla riga 1, colonne A a G
titles = []
for column in range(1, 2):
    cell_value = worksheet.cell(row=1, column=column).value
    titles.append(cell_value)

# Leggi i dati delle righe successive, dalle celle A2 a Gn (n = numero di righe)
data = []
for row in range(2, worksheet.max_row + 1):
    row_data = {}
    for column in range(1, 2):
        cell_value = worksheet.cell(row=row, column=column).value
        row_data[titles[column - 1]] = cell_value
    data.append(row_data)

# Chiudi il file Excel
workbook.close()

# Fai qualcosa con i dati letti...
print(titles)

COCOZZA_connectDb.start()



for record in data:
    if record['TARGA'] == None:
        continue
    PRINT("STAMPO RECORD")
    PRINT(record)
    targa = record['TARGA']
    PRINT(targa)
    
    # Recupero i valòori necessari per il recuper dei dati
    sconosciuto_brand_id = COCOZZA_connectDb.sqlSearch('fleet.vehicle.model.brand', ['name', '=', 'Sconosciuto'], '')[0]['id']
    sconosciuto_category_id = COCOZZA_connectDb.sqlSearch('fleet.vehicle.model.category', ['name', 'ilike', 'Sconosciuto'],'')[0]['id']
    sconosciuto_modello_id = COCOZZA_connectDb.sqlSearchMultiple('fleet.vehicle.model', [('name', '=', 'Sconosciuto'), ('category_id', '=', sconosciuto_category_id), ('brand_id', '=', sconosciuto_brand_id)])[0]['id']
    
    
    PRINT(sconosciuto_brand_id)
    PRINT(sconosciuto_category_id)
    PRINT(sconosciuto_modello_id)
    
    # Verifico se la targa è già inserita
    PRINT(f"Verifico se il mezzo targato {targa} è già inserito")
    is_targa = COCOZZA_connectDb.sqlSearch('fleet.vehicle', ['license_plate', '=', targa], '')    
    if is_targa == []:
        PRINT("Il mezzo non esiste. Procedo alla creazione del mezzo")
        COCOZZA_connectDb.sqlCreate("fleet.vehicle", {
                                                    'brand_id': sconosciuto_brand_id,
                                                    'license_plate': targa,
                                                    'model_id': sconosciuto_modello_id,
                                                    'stato_veicolo': 'RESTITUITO'
                                                    })
        PRINT("Veicolo aggiunto!")
    else:
        PRINT("Il mezzo è già inserito nel database")
    

PRINT("L'IMPORTAZIONE E' FINALMENTE TERMINATA!!!")